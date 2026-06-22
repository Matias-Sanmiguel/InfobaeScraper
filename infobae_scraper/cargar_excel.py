import argparse
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Side

THIN = Side(style="thin", color="D0D0D0")
BORDE = Border(left=THIN, right=THIN, bottom=THIN, top=THIN)
FUENTE_FIJA = "Infobae"


def leer_csv(directorio: Path, nombre: str) -> list[dict]:
    ruta = directorio / nombre
    if not ruta.exists():
        return []
    with open(ruta, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def construir_mapa(filas: list[dict], clave: str) -> dict[str, list[str]]:
    mapa: dict[str, list[str]] = {}
    for fila in filas:
        k = fila.get(clave, "")
        for v in fila.values():
            if v != k:
                mapa.setdefault(k, []).append(v)
    return mapa


def cargar_excel(directorio_csv: str, archivo_excel: str, archivo_salida: str):
    dir_csv = Path(directorio_csv)
    noticias = leer_csv(dir_csv, "noticias.csv")
    autores_filas = leer_csv(dir_csv, "autores.csv")
    temas_filas = leer_csv(dir_csv, "temas.csv")
    rel_autores = leer_csv(dir_csv, "rel_escrito_por.csv")
    rel_temas = leer_csv(dir_csv, "rel_menciona.csv")

    autores_por_id = {f["autor_id"]: f["nombre"] for f in autores_filas}
    temas_por_id = {f["tema_id"]: f["nombre"] for f in temas_filas}

    autores_por_noticia: dict[str, list[str]] = {}
    for rel in rel_autores:
        nid = rel["noticia_id"]
        autores_por_noticia.setdefault(nid, []).append(autores_por_id.get(rel["autor_id"], ""))

    temas_por_noticia: dict[str, list[str]] = {}
    for rel in rel_temas:
        nid = rel["noticia_id"]
        temas_por_noticia.setdefault(nid, []).append(temas_por_id.get(rel["tema_id"], ""))

    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active

    for noticia in noticias:
        nid = noticia["noticia_id"]
        autores = ", ".join(filter(None, autores_por_noticia.get(nid, [])))
        temas = ", ".join(filter(None, temas_por_noticia.get(nid, [])))

        fila = [
            noticia.get("noticia_id", ""),
            noticia.get("url", ""),
            noticia.get("titulo", ""),
            noticia.get("cuerpo", ""),
            noticia.get("fecha_publicacion", ""),
            noticia.get("fecha_modificacion", ""),
            noticia.get("seccion", ""),
            FUENTE_FIJA,
            autores,
            temas,
            noticia.get("imagen_portada", ""),
        ]

        fila_num = ws.max_row + 1
        for col, valor in enumerate(fila, 1):
            celda = ws.cell(row=fila_num, column=col, value=valor)
            celda.border = BORDE
            celda.alignment = Alignment(vertical="top", wrap_text=False)

    ruta_salida = Path(archivo_salida)
    wb.save(ruta_salida)
    print(f"OK: {len(noticias)} artículo(s) cargados en {ruta_salida}")


def main():
    parser = argparse.ArgumentParser(description="Carga CSVs del scraper al Excel de plantilla")
    parser.add_argument("--csvs", default="datos", help="Directorio con los CSVs (default: datos)")
    parser.add_argument("--plantilla", default="../plantilla_datos_infobae.xlsx", help="Archivo Excel plantilla")
    parser.add_argument("--salida", default="../datos_infobae.xlsx", help="Archivo Excel de salida")
    args = parser.parse_args()
    cargar_excel(args.csvs, args.plantilla, args.salida)


if __name__ == "__main__":
    main()
